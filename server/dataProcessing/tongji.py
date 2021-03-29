import json
import random
import math
from dataProcessing.Ged import getGed,getTed
import numpy as np
import matplotlib.pyplot as plt
import os
import re
import openpyxl as ox
from datetime import datetime

def getJson(url): #读取原始数据
    data = {}
    with open(url, 'r',encoding='utf-8') as fr:
        data=json.load(fr)

    return data

def getTreeNodes(tree):
    if 'id' in tree.keys():
        node = tree['id']
    elif 'name' in tree.keys():
        node = tree['name']
    nodes=[node]
    if 'children' in tree.keys():
        for i in range(len(tree['children'])):
            nodes.extend(getTreeNodes(tree['children'][i]))
    else:return nodes
    return nodes

def getTreeEdges(tree):
    edges=[]
    if 'id' in tree.keys():
        node=tree['id']
    elif 'name' in tree.keys():
        node=tree['name']
    if 'children' in tree.keys():
        for i in range(len(tree['children'])):
            edges.append([node,tree['children'][i]['name']])
            edges.extend(getTreeEdges(tree['children'][i]))
    else:return edges
    return edges


def run(dataType,dir,sampleNum,matchNum,dimensions,weight,dimensionsAttrChecked):

    graphs = getJson('./data/'+dataType+'/subGraphs_1.json')
    sampleNode = []
    sampleRandom=[]
    index = 0;
    while index < sampleNum:
        sample = random.randint(1, len(graphs)-1)
        if sample not in sampleRandom:
            sampleRandom.append(sample)
            sampleNode.append(graphs[sample]['id'])
            index += 1

    graphsEgdes = {}
    attrNamesAll = []
    if dataType=="Author":
        for i in graphs:
            graphsEgdes[str(i['id'])] = {'edges': i['edges'], 'attr': i['attr'], 'nodes': i['nodes']}
    else:
        with open('./data/'+dataType+'/subGraphs_tree.json','r') as fr:
            treeData=json.load(fr)
        for i in graphs:
            # edges=getTreeEdges(i)
            edges=treeData[str(i['id'])]
            # nodes=getTreeNodes(i)
            # if len(edges)==0:
            #     edges=[[nodes[0],nodes[0]]]
            graphsEgdes[str(i['id'])] = {'edges': edges, 'attr': i['attr']}
    attrIndex=list(i.start() for i in re.finditer('1', dimensionsAttrChecked))
    for i in graphs[0]['attr']:
        attrNamesAll.append(i)

    attrNames=[]
    for i in attrIndex:
        attrNames.append(attrNamesAll[i])
    print(attrNames)

    gedValueMeanEveryWeight = []  # 每个权重下ged均值
    gedValueStdEveryWeight = []  # 每个权重下ged方差
    attrEveryWeight = {}
    attrEveryWeightSum = []
    for name in attrNames:
        attrEveryWeight[name] = []
    for i in weight:
        attrEveryWeightSum.append(0)
        data = getJson('./data/'+dataType+'/vectors2d_1_' + str(dimensions) + '_' + str(i[0]) + '_' + str(i[1])+'_'+dimensionsAttrChecked+'.json')

        # 每个样本
        gedValueMeanEverySample = []  # 每个样本ged均值
        gedValueStdEverySample = []  # 每个样本ged方差
        attrEverySample = {}
        for name in attrNames:
            attrEverySample[name] = []
        for j in sampleNode:
            sourceX = float(data[str(j)]['x'])
            sourceY = float(data[str(j)]['y'])
            if dataType!='Family':
                sourceNodes = graphsEgdes[str(j)]['nodes']
            sourceEdges = graphsEgdes[str(j)]['edges']
            sourceAttr = graphsEgdes[str(j)]['attr']

            data2 = []
            for key in data:
                targetX = float(data[key]['x'])
                targetY = float(data[key]['y'])
                distance = math.sqrt(math.pow(targetX - sourceX, 2) + math.pow(targetY - sourceY, 2))
                data2.append({'id': key, 'x': targetX, 'y': targetY, 'distance': distance})
            matchGraphs = sorted(data2, key=lambda x: x['distance'])[1:matchNum + 1]
            # print(matchGraphs)
            gedValue = []
            attrValue = {}
            for name in attrNames:
                attrValue[name] = []

            for graph in matchGraphs:

                targetEdges = graphsEgdes[graph['id']]['edges']
                targetAttr = graphsEgdes[graph['id']]['attr']
                if dataType!='Family':
                    targetNodes = graphsEgdes[graph['id']]['nodes']
                    ged = getGed(sourceEdges, targetEdges, sourceNodes, targetNodes)
                else:
                    ged=getTed(sourceEdges,targetEdges)
                gedValue.append(ged)
                for name in attrNames:
                    attrValue[name].append(abs(targetAttr[name] - sourceAttr[name]))
            gedValueMeanEverySample.append(np.mean(np.array(gedValue)))
            gedValueStdEverySample.append(np.std(np.array(gedValue)))
            for name in attrNames:
                attrEverySample[name].append(np.mean(np.array(attrValue[name])))
        gedValueMeanEveryWeight.append(np.mean(np.array(gedValueMeanEverySample)))
        gedValueStdEveryWeight.append(np.std(np.array(gedValueStdEverySample)))
        for name in attrNames:
            attrEveryWeight[name].append(np.mean(np.array(attrEverySample[name])))

    # print(gedValueMeanEveryWeight)
    #
    # print(attrEveryWeight)

    dirname='+'.join(attrNames)

    for i in attrEveryWeight:
        attrEveryWeight[i]=translate(attrEveryWeight[i],0,1)
        for j in range(len(attrEveryWeight[i])):
            attrEveryWeightSum[j]+=attrEveryWeight[i][j]

    filepath='./data/'+dataType+'/images/'+dirname+'/sample'+str(sampleNum)+'match'+str(matchNum)+'/' + str(dir) + '/'
    if os.path.exists(filepath)==False:
        os.makedirs(filepath)
    y_data_mean = np.array(gedValueMeanEveryWeight)
    print(y_data_mean)
    # y_data_std = np.array(gedValueStdEveryWeight)
    x_data = ['Str','Attr','Str+Attr']
    # plt.legend(loc="upper right")  # 显示图例
    plt.xlabel('X')
    plt.ylabel('Ged Value')
    plt.bar(x_data, translate(y_data_mean,1,10),align='center')
    # plt.plot(x_data, y_data_std,color='blue', linewidth=2,label="Ged Std",marker = "o" ,markerfacecolor = 'Orange')
    plt.savefig('./data/'+dataType+'/images/'+dirname+'/sample'+str(sampleNum)+'match'+str(matchNum)+'/' + str(dir) + "/GedMean.png")
    plt.close()

    plt.xlabel('X')
    plt.ylabel(dirname)
    y_data_mean = np.array(attrEveryWeightSum)
    print(y_data_mean)
    plt.bar(x_data, translate(y_data_mean, 1, 10), align='center')
    plt.savefig('./data/' + dataType + '/images/' + dirname + '/sample'+str(sampleNum)+'match'+str(matchNum)+'/' + str(dir)+'/'+dirname + ".png")
    plt.close()

    if gedValueMeanEveryWeight[0]<=gedValueMeanEveryWeight[2] and gedValueMeanEveryWeight[2] <=gedValueMeanEveryWeight[1] and attrEveryWeightSum[1]<=attrEveryWeightSum[2] and attrEveryWeightSum[2]<=attrEveryWeightSum[0]:
        return True,'./data/'+dataType+'/images/'+dirname+'/sample'+str(sampleNum)+'match'+str(matchNum)+'/'
    else:
        return False,'./data/'+dataType+'/images/'+dirname+'/sample'+str(sampleNum)+'match'+str(matchNum)+'/'


def run2(dataType,dir,sampleNum,matchNum,dimensions,weight,dimensionsAttrChecked):

    graphs = getJson('./data/'+dataType+'/subGraphs_1.json')
    sampleNode = []
    sampleRandom=[]
    index = 0;
    while index < sampleNum:
        sample = random.randint(1, len(graphs)-1)
        if sample not in sampleRandom:
            sampleRandom.append(sample)
            sampleNode.append(graphs[sample]['id'])
            index += 1

    graphsEgdes = {}
    attrNamesAll = []
    if dataType=="Author":
        for i in graphs:
            graphsEgdes[str(i['id'])] = {'edges': i['edges'], 'attr': i['attr'], 'nodes': i['nodes']}
    else:
        for i in graphs:
            edges=getTreeEdges(i)
            nodes=getTreeNodes(i)
            if len(edges)==0:
                edges=[[nodes[0],nodes[0]]]
            graphsEgdes[str(i['id'])] = {'edges': edges, 'attr': i['attr'], 'nodes': nodes}
    attrIndex=list(i.start() for i in re.finditer('1', dimensionsAttrChecked))
    for i in graphs[0]['attr']:
        attrNamesAll.append(i)

    attrNames=[]
    for i in attrIndex:
        attrNames.append(attrNamesAll[i])
    print(attrNames)

    gedValueMeanEveryWeight = []  # 每个权重下ged均值
    gedValueStdEveryWeight = []  # 每个权重下ged方差
    attrEveryWeight = {}
    attrEveryWeightSum = []
    for name in attrNames:
        attrEveryWeight[name] = []
    for i in weight:
        attrEveryWeightSum.append(0)
        data = getJson('./data/'+dataType+'/vectors2d_1_' + str(dimensions) + '_' + str(i[0]) + '_' + str(i[1])+'_'+dimensionsAttrChecked+'.json')

        # 每个样本
        gedValueMeanEverySample = []  # 每个样本ged均值
        gedValueStdEverySample = []  # 每个样本ged方差
        attrEverySample = {}
        for name in attrNames:
            attrEverySample[name] = []
        for j in sampleNode:
            sourceX = float(data[str(j)]['x'])
            sourceY = float(data[str(j)]['y'])
            sourceEdges = graphsEgdes[str(j)]['edges']
            sourceNodes = graphsEgdes[str(j)]['nodes']
            sourceAttr = graphsEgdes[str(j)]['attr']

            data2 = []
            for key in data:
                targetX = float(data[key]['x'])
                targetY = float(data[key]['y'])
                distance = math.sqrt(math.pow(targetX - sourceX, 2) + math.pow(targetY - sourceY, 2))
                data2.append({'id': key, 'x': targetX, 'y': targetY, 'distance': distance})
            matchGraphs = sorted(data2, key=lambda x: x['distance'])[1:matchNum + 1]
            # print(matchGraphs)
            gedValue = []
            attrValue = {}
            for name in attrNames:
                attrValue[name] = []

            for graph in matchGraphs:
                targetEdges = graphsEgdes[graph['id']]['edges']
                targetNodes = graphsEgdes[graph['id']]['nodes']
                targetAttr = graphsEgdes[graph['id']]['attr']
                ged = getGed(sourceEdges, targetEdges, sourceNodes, targetNodes)
                gedValue.append(ged)
                for name in attrNames:
                    attrValue[name].append(abs(targetAttr[name] - sourceAttr[name]))
            gedValueMeanEverySample.append(np.mean(np.array(gedValue)))
            gedValueStdEverySample.append(np.std(np.array(gedValue)))
            for name in attrNames:
                attrEverySample[name].append(np.mean(np.array(attrValue[name])))
        gedValueMeanEveryWeight.append(np.mean(np.array(gedValueMeanEverySample)))
        gedValueStdEveryWeight.append(np.std(np.array(gedValueStdEverySample)))
        for name in attrNames:
            attrEveryWeight[name].append(np.mean(np.array(attrEverySample[name])))

    # print(gedValueMeanEveryWeight)
    #
    # print(attrEveryWeight)
    for attr in attrEveryWeight:
        attrEveryWeight[attr]=translate(attrEveryWeight[attr],0.1,1)
    gedValueMeanEveryWeight=translate(gedValueMeanEveryWeight,0.1,1)
    dirname='+'.join(attrNames)
    ways = ['Str', 'Attr', 'Str+Attr']
    data=[]
    for i in range(len(ways)):
        typeData=[]
        for attr in attrEveryWeight:
            typeData.append(attrEveryWeight[attr][i])
        typeData.append(gedValueMeanEveryWeight[i])
        data.append(typeData)
    attrNames.append("GED")

    x = np.arange(len(attrNames))  # the label locations
    width = 0.8/3  # the width of the bars

    fig, ax = plt.subplots()

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('Value')
    ax.set_title('Scores by different ways')
    ax.set_xticks(x)
    ax.set_xticklabels(attrNames)

    for i in range(len(ways)):
        ax.bar(x + width*(i-1), data[i], width, label=ways[i])
    ax.legend(loc="upper right")

    fig.tight_layout()
    filepath='./data/'+dataType+'/images/'+dirname+'/sample'+str(sampleNum)+'match'+str(matchNum)+'/'
    if os.path.exists(filepath)==False:
        os.makedirs(filepath)
    plt.savefig(filepath+ str(dir) + ".png")
    plt.close()

    for i in range(len(attrNames)):
        if i!=len(attrNames)-1:
            if data[0][i]<data[1][i] or data[0][i]<data[2][i] or data[2][i]<data[1][i]:
                return False,filepath
        else:
            if data[0][i]>data[1][i] or data[0][i]>data[2][i] or data[2][i]>data[1][i]:
                return False, filepath
    return True, filepath

def run1(dataType, sampleNum, matchNum, dimensions, weight, dimensionsAttrChecked):

    graphs = getJson('./data/' + dataType + '/subGraphs_1.json')

    sampleNode = []
    sampleRandom = []
    index = 0;
    while index < sampleNum:
        sample = random.randint(1, len(graphs) - 1)
        if sample not in sampleRandom:
            sampleRandom.append(sample)
            sampleNode.append(graphs[sample]['id'])
            index += 1

    graphsEgdes = {}
    attrNamesAll = []
    if dataType == "Author":
        for i in graphs:
            graphsEgdes[str(i['id'])] = {'edges': i['edges'], 'attr': i['attr'], 'nodes': i['nodes']}
    else:
        with open('./data/'+dataType+'/subGraphs_tree.json','r') as fr:
            treeData=json.load(fr)
        for i in graphs:
            edges = treeData[str(i['id'])]
            # nodes = getTreeNodes(i)
            # if len(edges) == 0:
            #     edges = [[nodes[0], nodes[0]]]
            graphsEgdes[str(i['id'])] = {'edges': edges, 'attr': i['attr']}
    attrIndex = list(i.start() for i in re.finditer('1', dimensionsAttrChecked))
    for i in graphs[0]['attr']:
        attrNamesAll.append(i)

    attrNames = []
    for i in attrIndex:
        attrNames.append(attrNamesAll[i])
    print(attrNames)

    gedValueMeanEveryWeight = []  # 每个权重下ged均值
    gedValueStdEveryWeight = []  # 每个权重下ged方差
    attrEveryWeight = {}
    attrEveryWeightSum = []
    for name in attrNames:
        attrEveryWeight[name] = []
    for i in weight:
        attrEveryWeightSum.append(0)
        if i[0]==3:
            dimensions = dimensionsAttrChecked.count('1')
        data = getJson('./data/' + dataType + '/vectors2d_1_' + str(dimensions) + '_' + str(i[0]) + '_' + str(
            i[1]) + '_' + dimensionsAttrChecked + '.json')

        # 每个样本
        gedValueMeanEverySample = []  # 每个样本ged均值
        gedValueStdEverySample = []  # 每个样本ged方差
        attrEverySample = {}
        for name in attrNames:
            attrEverySample[name] = []
        for j in sampleNode:
            sourceX = float(data[str(j)]['x'])
            sourceY = float(data[str(j)]['y'])
            sourceAttr = graphsEgdes[str(j)]['attr']
            sourceEdges = graphsEgdes[str(j)]['edges']
            if dataType!='Family':
                sourceNodes = graphsEgdes[str(j)]['nodes']
            data2 = []
            for key in data:
                targetX = float(data[key]['x'])
                targetY = float(data[key]['y'])
                distance = math.sqrt(math.pow(targetX - sourceX, 2) + math.pow(targetY - sourceY, 2))
                data2.append({'id': key, 'x': targetX, 'y': targetY, 'distance': distance})
            matchGraphs = sorted(data2, key=lambda x: x['distance'])[1:matchNum + 1]
            # print(matchGraphs)
            gedValue = []
            attrValue = {}
            for name in attrNames:
                attrValue[name] = []

            for graph in matchGraphs:
                targetAttr = graphsEgdes[graph['id']]['attr']
                targetEdges = graphsEgdes[graph['id']]['edges']
                if dataType!='Family':
                    targetNodes = graphsEgdes[graph['id']]['nodes']
                    ged = getGed(sourceEdges, targetEdges, sourceNodes, targetNodes)
                else:
                    ged=getTed(sourceEdges,targetEdges)
                gedValue.append(ged)
                for name in attrNames:
                    attrValue[name].append(abs(targetAttr[name] - sourceAttr[name]))
            gedValueMeanEverySample.append(np.sum(np.array(gedValue)))
            gedValueStdEverySample.append(np.std(np.array(gedValue)))
            for name in attrNames:
                attrEverySample[name].append(np.sum(np.array(attrValue[name])))
        gedValueMeanEveryWeight.append(np.sum(np.array(gedValueMeanEverySample)))
        gedValueStdEveryWeight.append(np.std(np.array(gedValueStdEverySample)))
        for name in attrNames:
            attrEveryWeight[name].append(np.sum(np.array(attrEverySample[name])))

    dirname = '+'.join(attrNames)

    for i in attrEveryWeight:
        for j in range(len(attrEveryWeight[i])):
            attrEveryWeightSum[j] += attrEveryWeight[i][j]

    print(gedValueMeanEveryWeight)

    print(attrEveryWeightSum)

    return {'str':gedValueMeanEveryWeight,'attr':attrEveryWeightSum}



def translate(array,rightMin, rightMax):
    # Figure out how 'wide' each range is
    leftMax=max(array)
    leftMin=min(array)
    leftSpan = leftMax - leftMin
    rightSpan = rightMax - rightMin

    # Convert the left range into a 0-1 range (float)
    newArray=[]
    for i in array:
        valueScaled = float(i - leftMin) / float(leftSpan)
        newArray.append(rightMin + (valueScaled * rightSpan))

    # Convert the 0-1 range into a value in the right range.
    return newArray

def createExcel(data,path,columnNames,matchNodes):
    wb=ox.Workbook()

    ws=wb.worksheets[0]
    columnNames.insert(0, "type")
    columnNames.insert(0, "matchNum")
    ws.append(columnNames)
    step=2
    for i in range(len(matchNum)):
        index="A"+str(i+step)+":A"+str(i+step+1)
        ws.merge_cells(index)
        ws["A"+str(i+step)]=matchNum[i]

        data[i]['str'].insert(0, "Str")
        data[i]['attr'].insert(0, "Attr")
        for index in range(2,8):
            ws.cell(column=index,row=i+step,value=data[i]['str'][index-2])
            ws.cell(column=index, row=i + step+1, value=data[i]['attr'][index - 2])
        step += 1

    wb.save(path)



if __name__=="__main__":
    start=datetime.now()
    dataType="Family"
    dir=1
    sampleNum = [50]
    matchNum = [20,40,60]
    dimensions = 128
    dimensionsAttrChecked = ['11111']
    weight = [[1, 0],[0,1],[1,1],[2,1],[3,1]]
    for p in range(len(dimensionsAttrChecked)):
        for j in range(len(sampleNum)):
            excelData=[]
            for k in range(len(matchNum)):
                results=[]
                for i in range(dir):
                    print("sampleNum:", sampleNum[j], "matchNum:", matchNum[k],"第"+str(i)+"次")
                    results.append(run1(dataType, sampleNum[j], matchNum[k], dimensions, weight, dimensionsAttrChecked[p]))
                data=results[0]
                for index in range(1,len(results)):
                    for name in results[index]:
                        for col in range(len(results[index][name])):
                            data[name][col]+=results[index][name][col]
                for name in data:
                    for i in range(len(data[name])):
                        data[name][i]/=dir
                excelData.append(data)
            print(excelData)
            createExcel(excelData,'./data/'+dataType+'/tongji.xlsx',['str','attr','cca','str128+attr','str=attr'],matchNum)
                #     result,path=run2(dataType, i, sampleNum[j], matchNum[k], dimensions, weight, dimensionsAttrChecked[p])
                #     results.append(result)
                # with open(path+'result2.txt','w',encoding='utf-8') as fr:
                #     fr.write("准确率："+str(results.count(True)/len(results)*100)+"%")
                # print("准确率："+str(results.count(True)/len(results)*100)+"%")
    end = datetime.now()
    print((end - start).seconds)